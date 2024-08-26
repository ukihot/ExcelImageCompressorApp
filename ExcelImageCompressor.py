import os
import time
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
from io import BytesIO
from datetime import datetime
import subprocess
import sys


# フォルダーを選択するためのダイアログを表示する関数
def select_folder():
    return filedialog.askdirectory()


# 指定したフォルダー内のExcelファイル（.xlsx、.xlsm）を検索する関数
def find_excel_files(folder):
    return [
        os.path.join(root, file)
        for root, _, files in os.walk(folder)
        for file in files
        if (file.endswith((".xlsx", ".xlsm")) and not file.startswith("~$"))
    ]


# 画像を圧縮する関数
def compress_image(image):
    img_data = image._data()  # 画像データを取得
    with Image.open(BytesIO(img_data)) as img:  # 画像データを開く
        img_byte_arr = BytesIO()  # バイト配列を作成
        img.save(
            img_byte_arr, format=img.format if img.format else "PNG", dpi=(96, 96)
        )  # 画像を圧縮して保存
        return OpenpyxlImage(img_byte_arr)  # 圧縮後の画像を返す


# ファイルサイズをKB単位で取得する関数
def get_file_size_in_kb(file_path):
    return os.path.getsize(file_path) / 1024


# Excelファイル内の画像を圧縮する関数
def compress_images_in_file(file):
    start_time = time.time()  # 処理開始時刻を記録
    original_size = get_file_size_in_kb(file)  # 圧縮前のファイルサイズを取得
    error_message = None  # エラーメッセージを初期化

    try:
        workbook = load_workbook(file)  # Excelファイルを読み込み
        total_images = sum(
            len(worksheet._images) for worksheet in workbook.worksheets
        )  # 画像の総数を取得
        updated_images = 0  # 更新した画像のカウント

        for worksheet in workbook.worksheets:
            try:
                worksheet._images[:] = [
                    compress_image(image) for image in worksheet._images  # 画像を圧縮
                ]
                updated_images += len(
                    worksheet._images
                )  # 更新した画像のカウントを増やす
            except Exception as e:
                error_message = f"{str(e)}"  # エラーメッセージを記録
                break  # エラーが発生した場合は処理を中断

        if not error_message and updated_images > 0:
            workbook.save(file)  # 圧縮後のExcelファイルを保存

    except Exception as e:
        error_message = str(e)  # エラーメッセージを記録

    end_time = time.time()  # 処理終了時刻を記録
    new_size = (
        get_file_size_in_kb(file) if not error_message else original_size
    )  # 圧縮後のファイルサイズを取得
    elapsed_time = (end_time - start_time) * 1000  # 処理時間をミリ秒単位で計算
    return original_size, new_size, updated_images, elapsed_time, error_message


# 圧縮結果を報告するExcelファイルを作成する関数
def create_report(report_data, folder):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # タイムスタンプを作成
    report_file = os.path.join(
        folder, f"report_{timestamp}.xlsx"
    )  # 報告書ファイルのパスを作成
    workbook = Workbook()  # 新しいExcelファイルを作成
    sheet = workbook.active  # アクティブなシートを取得
    sheet.title = "Report"  # シートのタイトルを設定

    headers = [
        "Date",
        "Orig Size (KB)",
        "New Size (KB)",
        "Img Count",
        "Path",
        "File",
        "Time (ms)",
        "Status",
        "Error Message",
    ]
    sheet.append(headers)  # ヘッダー行を追加

    for row in report_data:
        # パスとファイル名が同じ場合はパスを'.'に設定
        path, file = row[4], row[5]
        if path == file:
            row[4] = "."

        sheet.append(
            [
                f"{value:.1f}" if isinstance(value, (int, float)) else value
                for value in row
            ]
        )  # データ行を追加

    # 各列のフォーマットを設定
    for col in range(1, 10):
        for cell in sheet.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                if isinstance(c.value, (int, float)):
                    c.number_format = "0.0"

    workbook.save(report_file)  # 報告書ファイルを保存
    return report_file


# フォルダー内のすべてのExcelファイルの画像を圧縮し、報告書を作成する関数
def compress_images_in_folder(folder, progress_callback):
    excel_files = find_excel_files(folder)  # フォルダー内のExcelファイルを取得
    report_data = []  # 報告書用データのリストを初期化

    for i, file in enumerate(excel_files, start=1):
        original_size, new_size, updated_images, elapsed_time, error_message = (
            compress_images_in_file(file)  # ファイルごとに画像を圧縮
        )
        relative_path = os.path.relpath(file, folder)  # 相対パスを取得
        file_name = os.path.basename(file)  # ファイル名を取得
        processing_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 処理日付を取得

        status = "Success" if not error_message else "Failed"  # ステータスを設定
        report_data.append(
            [
                processing_date,
                original_size,
                new_size,
                updated_images,
                relative_path,
                file_name,
                elapsed_time,
                status,
                error_message or "",
            ]
        )  # データ行を追加

        progress_callback(i, len(excel_files))  # 進行状況を更新

    report_file = create_report(report_data, folder)  # 報告書を作成
    return (
        len(excel_files),
        report_file,
    )  # 処理したファイル数と報告書ファイルのパスを返す


# GUIアプリケーションのクラス
class ExcelImageCompressor:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Image Compressor")  # ウィンドウタイトルを設定
        self.selected_folder = (
            tk.StringVar()
        )  # 選択されたフォルダーのパスを保持する変数

        # UIの設定
        frame = tk.Frame(root, padx=10, pady=10)
        frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.folder_label = tk.Label(frame, text="Select Folder:")
        self.folder_label.grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)

        self.folder_entry = tk.Entry(frame, textvariable=self.selected_folder, width=40)
        self.folder_entry.grid(row=0, column=1, padx=5, pady=5)

        self.select_button = tk.Button(frame, text="Browse", command=self.set_folder)
        self.select_button.grid(row=0, column=2, padx=5, pady=5)

        self.start_button = tk.Button(
            frame, text="Start Compression", command=self.compress_images
        )
        self.start_button.grid(row=1, column=0, columnspan=3, pady=10)

        self.progress = Progressbar(
            frame, orient="horizontal", length=300, mode="determinate"
        )
        self.progress.grid(row=2, column=0, columnspan=2, pady=10)

        self.progress_label = tk.Label(frame, text="0.00%")
        self.progress_label.grid(row=2, column=2, padx=5, pady=10, sticky=tk.W)

    # フォルダー選択ボタンの処理
    def set_folder(self):
        self.selected_folder.set(select_folder())

    # 圧縮処理を開始するメソッド
    def compress_images(self):
        folder = self.selected_folder.get()
        if not folder:
            messagebox.showerror(
                "Error", "No folder selected."
            )  # フォルダーが選択されていない場合のエラーメッセージ
            return

        try:
            total_files, report_file = compress_images_in_folder(
                folder, self.update_progress
            )
            messagebox.showinfo(
                "Finished",
                f"Compression completed for {total_files} files. Report saved to:\n{report_file}",
            )  # 圧縮完了のメッセージ
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")  # エラーメッセージ

    # 進行状況を更新するメソッド
    def update_progress(self, current, total):
        percentage = (current / total) * 100
        self.progress["value"] = percentage
        self.progress_label.config(text=f"{percentage:.2f}%")
        self.root.update_idletasks()


# メイン関数
def main():
    root = tk.Tk()  # Tkinterのルートウィンドウを作成
    app = ExcelImageCompressor(root)  # アプリケーションのインスタンスを作成
    root.mainloop()  # イベントループを開始


if __name__ == "__main__":
    main()  # メイン関数を呼び出してアプリケーションを実行
