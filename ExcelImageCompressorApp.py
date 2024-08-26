import os
import time
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
from io import BytesIO
from datetime import datetime
import subprocess
import sys


def select_folder():
    return filedialog.askdirectory()


def find_excel_files(folder):
    return [
        os.path.join(root, file)
        for root, _, files in os.walk(folder)
        for file in files
        if (file.endswith((".xlsx", ".xlsm")) and not file.startswith("~$"))
    ]


def compress_image(image):
    img_data = image._data()
    with Image.open(BytesIO(img_data)) as img:
        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format=img.format if img.format else "PNG", dpi=(96, 96))
        return OpenpyxlImage(img_byte_arr)


def get_file_size_in_kb(file_path):
    return os.path.getsize(file_path) / 1024


def compress_images_in_file(file):
    start_time = time.time()
    original_size = get_file_size_in_kb(file)
    error_message = None

    try:
        workbook = load_workbook(file)
        total_images = sum(len(worksheet._images) for worksheet in workbook.worksheets)
        updated_images = 0

        for worksheet in workbook.worksheets:
            try:
                worksheet._images[:] = [
                    compress_image(image) for image in worksheet._images
                ]
                updated_images += len(worksheet._images)
            except Exception as e:
                error_message = f"{str(e)}"
                break

        if not error_message and updated_images > 0:
            workbook.save(file)

    except Exception as e:
        error_message = str(e)

    end_time = time.time()
    new_size = get_file_size_in_kb(file) if not error_message else original_size
    elapsed_time = (end_time - start_time) * 1000  # in milliseconds
    return original_size, new_size, updated_images, elapsed_time, error_message


def create_report(report_data, folder):
    report_file = os.path.join(folder, "report.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    headers = [
        "Date",
        "Orig Size (KB)",
        "New Size (KB)",
        "Img Count",
        "Path",
        "File",
        "Time (ms)",
        "Status",
        "Error Message",
    ]
    sheet.append(headers)

    for row in report_data:
        sheet.append(
            [
                f"{value:.1f}" if isinstance(value, (int, float)) else value
                for value in row
            ]
        )

    for col in range(1, 10):
        for cell in sheet.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                if isinstance(c.value, (int, float)):
                    c.number_format = "0.0"

    workbook.save(report_file)
    return report_file


def compress_images_in_folder(folder, progress_callback):
    excel_files = find_excel_files(folder)
    report_data = []

    for i, file in enumerate(excel_files, start=1):
        original_size, new_size, updated_images, elapsed_time, error_message = (
            compress_images_in_file(file)
        )
        relative_path = os.path.relpath(file, folder)
        file_name = os.path.basename(file)
        processing_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        status = "Success" if not error_message else "Failed"
        report_data.append(
            [
                processing_date,
                original_size,
                new_size,
                updated_images,
                relative_path,
                file_name,
                elapsed_time,
                status,
                error_message or "",
            ]
        )

        progress_callback(i, len(excel_files))

    report_file = create_report(report_data, folder)
    return len(excel_files), report_file


class ExcelImageCompressorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Image Compressor")
        self.selected_folder = tk.StringVar()

        # UI Setup
        frame = tk.Frame(root, padx=10, pady=10)
        frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.folder_label = tk.Label(frame, text="Select Folder:")
        self.folder_label.grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)

        self.folder_entry = tk.Entry(frame, textvariable=self.selected_folder, width=40)
        self.folder_entry.grid(row=0, column=1, padx=5, pady=5)

        self.select_button = tk.Button(frame, text="Browse", command=self.set_folder)
        self.select_button.grid(row=0, column=2, padx=5, pady=5)

        self.start_button = tk.Button(
            frame, text="Start Compression", command=self.compress_images
        )
        self.start_button.grid(row=1, column=0, columnspan=3, pady=10)

        self.progress = Progressbar(
            frame, orient="horizontal", length=300, mode="determinate"
        )
        self.progress.grid(row=2, column=0, columnspan=3, pady=10)

    def set_folder(self):
        self.selected_folder.set(select_folder())

    def compress_images(self):
        folder = self.selected_folder.get()
        if not folder:
            messagebox.showerror("Error", "No folder selected.")
            return

        try:
            total_files, report_file = compress_images_in_folder(
                folder, self.update_progress
            )
            messagebox.showinfo(
                "Completed",
                f"Compression completed for {total_files} files. Report generated in {folder}.",
            )
            os.startfile(report_file)  # Open the report file
            self.root.destroy()  # Close the application
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def update_progress(self, current, total):
        self.progress["maximum"] = total
        self.progress["value"] = current
        self.root.update_idletasks()


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelImageCompressorApp(root)
    root.mainloop()
